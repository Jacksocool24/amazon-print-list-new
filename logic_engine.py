"""
Data processing, image indexing, and Excel export.
"""
import io
import re
import base64
import pandas as pd
from PIL import Image, ImageStat
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def excel_text_str(val):
    """Normalize cell text; avoid scientific notation for long numbers."""
    if val is None or pd.isna(val):
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def qty_gt_one(val):
    try:
        return float(str(val).strip()) > 1
    except Exception:
        return False


def build_image_index(uploaded_images):
    """Index uploaded images by filename stem (lowercase)."""
    image_map = {}
    allowed_ext = {".jpg", ".jpeg", ".png"}
    for img in uploaded_images or []:
        file_name = (img.name or "").strip().replace("\\", "/").split("/")[-1]
        if not file_name:
            continue
        lower_name = file_name.lower()
        dot_idx = lower_name.rfind(".")
        if dot_idx == -1:
            continue
        ext = lower_name[dot_idx:]
        if ext not in allowed_ext:
            continue
        key = lower_name[:dot_idx]
        image_map[key] = {"bytes": img.getvalue(), "ext": ext}
    return image_map


def extract_size_parts(size_val):
    try:
        size_num = int(float(str(size_val).strip()))
    except Exception:
        return None, None
    digits = str(abs(size_num))
    if len(digits) == 4:
        return int(digits[:2]), int(digits[2:])
    if len(digits) == 3:
        return int(digits[:1]), int(digits[1:])
    if len(digits) >= 2:
        mid = len(digits) // 2
        return int(digits[:mid]), int(digits[mid:])
    if len(digits) == 1:
        val = int(digits)
        return val, val
    return None, None


def resolve_row_image_bytes(row, image_map):
    img_name_key = excel_text_str(row.get("图片名称", "")).lower()
    if img_name_key and img_name_key in image_map and image_map[img_name_key].get("bytes"):
        return image_map[img_name_key]["bytes"]
    order_key = f"{excel_text_str(row.get('purchase-date', ''))}_{excel_text_str(row.get('运单号', ''))}".lower()
    if order_key and order_key in image_map and image_map[order_key].get("bytes"):
        return image_map[order_key]["bytes"]
    return None


def detect_hanger_and_size(img_bytes, size_val):
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        img = img.resize((400, 400))
    except Exception:
        return "Canvas"

    width, height = img.size
    orientation = "portrait" if height >= width else "landscape"

    top_h = min(15, height)
    top_w = min(400, width)
    if top_h <= 0 or top_w <= 0:
        return "Canvas"
    top_region = img.crop((0, 0, top_w, top_h))
    top_std = ImageStat.Stat(top_region).stddev
    mean_std = sum(top_std) / len(top_std) if top_std else 0
    if mean_std < 8:
        return "Canvas"

    band_y0 = min(15, height)
    band_y1 = min(25, height)
    if band_y1 <= band_y0:
        return "Canvas"
    color_region = img.crop((0, band_y0, top_w, band_y1))
    avg = ImageStat.Stat(color_region).mean
    if len(avg) < 3:
        return "Canvas"
    r, g, b = avg[:3]
    hanger_type = "Canvas"
    if r < 60 and g < 60 and b < 60:
        hanger_type = "Black Hanger"
    elif r > 150 and g > 120 and b < 100:
        hanger_type = "Wood Hanger"
    if hanger_type == "Canvas":
        return "Canvas"

    first_len, second_len = extract_size_parts(size_val)
    if first_len is None or second_len is None:
        return hanger_type
    selected_len = first_len if orientation == "portrait" else second_len
    return f'{hanger_type} - {selected_len}"'


def apply_material_recognition(df, image_map):
    if df.empty:
        return df
    updated_df = df.copy()
    for i, row in updated_df.iterrows():
        img_bytes = resolve_row_image_bytes(row, image_map)
        updated_df.at[i, "材质"] = detect_hanger_and_size(img_bytes, row.get("尺寸")) if img_bytes else "Canvas"
    updated_df["图片名称"] = (
        updated_df["序号E"].astype(str) + "-" + updated_df["材质"] + "-" + updated_df["尺寸"].astype(str)
    )
    return updated_df


def process_data(source_df, ref_df):
    target = pd.DataFrame()
    target["运单号"] = source_df["order-item-id"].map(excel_text_str)
    target["purchase-date"] = source_df["order-id"].map(excel_text_str)
    target["SKU"] = source_df["sku"]
    target["图片"] = source_df.iloc[:, 2]
    target["数量"] = source_df["quantity-purchased"]
    target["姓名"] = source_df["recipient-name"]
    target["地址一"] = source_df["ship-address-1"]
    target["地址二"] = source_df["ship-address-2"].fillna("0")
    target["城市"] = source_df["ship-city"]
    target["州"] = source_df["ship-state"]
    target["邮编"] = source_df["ship-postal-code"]
    target["电话"] = source_df["ship-phone-number"].map(excel_text_str)
    target["Original Row Index"] = source_df.index + 2

    size_dict = dict(zip(ref_df["SKU"].astype(str), ref_df["尺寸"]))
    target["尺寸_原始"] = target["SKU"].astype(str).map(size_dict)
    target["尺寸_数值"] = pd.to_numeric(target["尺寸_原始"], errors="coerce").fillna(0).astype(int)
    target["尺寸"] = target["尺寸_数值"]

    target["Identity"] = target["姓名"].astype(str) + target["地址一"].astype(str)

    package_info = target.groupby("Identity").agg(
        row_count=("SKU", "count"),
        min_size=("尺寸_数值", "min"),
        max_size=("尺寸_数值", "max"),
        is_mixed_size=("尺寸_数值", lambda x: x.nunique() > 1),
    ).reset_index()

    def categorize(row):
        if row["is_mixed_size"]:
            return "B2"
        if row["row_count"] == 1 and row["max_size"] > 2436:
            return "A2"
        if row["max_size"] <= 2436:
            return "A1B1"
        return "A2"

    package_info["Category"] = package_info.apply(categorize, axis=1)
    target = target.merge(package_info[["Identity", "Category", "min_size", "max_size"]], on="Identity")

    tier1 = target[target["Category"] == "A1B1"].sort_values(
        ["max_size", "Identity", "尺寸_数值"], ascending=[False, True, False]
    )
    tier2 = target[target["Category"] == "B2"].sort_values(
        ["min_size", "Identity", "尺寸_数值"], ascending=[True, True, True]
    )
    tier3 = target[target["Category"] == "A2"].sort_values(
        ["尺寸_数值", "Identity"], ascending=[True, True]
    )

    final_df = pd.concat([tier1, tier2, tier3], ignore_index=True)
    final_df = final_df.merge(package_info[["Identity", "row_count"]], on="Identity", how="left")

    final_df["材质"] = "Canvas"
    final_df["材质"] = final_df["材质"].replace("画芯", "Canvas")

    current_idx = 0
    last_id = None
    d_column = []
    e_column = []
    id_counts = {}

    for _, row in final_df.iterrows():
        curr_id = row["Identity"]
        n_rows = int(row["row_count"])
        if curr_id != last_id:
            current_idx += 1
            id_counts[curr_id] = 1
            d_column.append(current_idx)
            if n_rows > 1:
                e_column.append(f"{current_idx}-1")
            else:
                e_column.append(current_idx)
        else:
            id_counts[curr_id] += 1
            d_column.append(current_idx)
            e_column.append(f"{current_idx}-{id_counts[curr_id]}")
        last_id = curr_id

    final_df["序号D"] = d_column
    final_df["序号E"] = e_column
    final_df["图片名称"] = final_df["序号E"].astype(str) + "-" + final_df["材质"] + "-" + final_df["尺寸"].astype(str)

    col_order = [
        "运单号",
        "purchase-date",
        "SKU",
        "序号D",
        "序号E",
        "图片",
        "材质",
        "尺寸",
        "数量",
        "姓名",
        "地址一",
        "地址二",
        "城市",
        "州",
        "邮编",
        "电话",
        "图片名称",
    ]
    return final_df[col_order + ["Original Row Index"]]


def save_to_excel_with_merge(df, source_file_bytes, image_map):
    output = io.BytesIO()
    wb = load_workbook(io.BytesIO(source_file_bytes), data_only=False, keep_vba=True)
    source_ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    source_sheet_name = source_ws.title

    if "最终打印清单" in wb.sheetnames:
        del wb["最终打印清单"]
    ws = wb.create_sheet("最终打印清单")

    text_cols = {1, 2, 16}
    img_col = 6
    output_cols = [
        "运单号",
        "purchase-date",
        "SKU",
        "序号D",
        "序号E",
        "图片",
        "材质",
        "尺寸",
        "数量",
        "姓名",
        "地址一",
        "地址二",
        "城市",
        "州",
        "邮编",
        "电话",
        "图片名称",
    ]

    ws.append(output_cols)

    ws.column_dimensions["F"].width = 28
    source_ws.column_dimensions["C"].width = 35
    row_height_pt = 240
    body_font = Font(name="SimSun", size=11, bold=True)
    header_font = Font(name="SimSun", size=12, bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    base_widths = {
        1: 20,
        2: 20,
        3: 16,
        4: 8,
        5: 10,
        6: 28,
        7: 10,
        8: 8,
        9: 8,
        10: 16,
        11: 24,
        12: 24,
        13: 12,
        14: 10,
        15: 12,
        16: 18,
        17: 22,
    }
    core_cols = {1, 2, 4, 5, 6, 16}
    for col_idx, width in base_widths.items():
        if col_idx == 6:
            ws.column_dimensions["F"].width = 28
            continue
        final_width = width if col_idx in core_cols else round(width * 0.9, 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = final_width

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        source_row = (
            int(row["Original Row Index"])
            if "Original Row Index" in row and pd.notna(row["Original Row Index"])
            else r_idx
        )
        img_key = f"{excel_text_str(row['purchase-date'])}_{excel_text_str(row['运单号'])}".lower()
        img_info = image_map.get(img_key)
        if img_info:
            img_stream = io.BytesIO(img_info["bytes"])
            xl_img = XLImage(img_stream)
            source_ws.row_dimensions[source_row].height = row_height_pt
            max_w, max_h = 236, 306
            w, h = float(xl_img.width), float(xl_img.height)
            if w > 0 and h > 0:
                scale = min(max_w / w, max_h / h, 1.0)
                xl_img.width = int(w * scale)
                xl_img.height = int(h * scale)
            source_ws.add_image(xl_img, f"C{source_row}")
            try:
                if hasattr(xl_img.anchor, "_from"):
                    x_offset = max(0, int((250 - xl_img.width) / 2))
                    y_offset = max(0, int((320 - xl_img.height) / 2))
                    xl_img.anchor._from.colOff = int(x_offset * 9525)
                    xl_img.anchor._from.rowOff = int(y_offset * 9525)
            except Exception:
                pass

    for row_idx in range(2, source_ws.max_row + 1):
        for col_idx in text_cols:
            src_cell = source_ws.cell(row=row_idx, column=col_idx)
            src_cell.value = excel_text_str(src_cell.value)
            src_cell.number_format = "@"

    safe_source_sheet = source_sheet_name.replace("'", "''")
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        source_row = (
            int(row["Original Row Index"])
            if "Original Row Index" in row and pd.notna(row["Original Row Index"])
            else r_idx
        )
        ws.row_dimensions[r_idx].height = row_height_pt
        for c_idx, col_name in enumerate(output_cols, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            value = row[col_name]
            if c_idx in text_cols:
                cell.value = excel_text_str(value)
                cell.number_format = "@"
            elif c_idx == img_col:
                cell.value = f"='{safe_source_sheet}'!C{source_row}"
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = body_font
            cell.border = thin_border
            if c_idx == 9 and qty_gt_one(value):
                cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
                cell.font = Font(name=body_font.name, size=body_font.size, bold=True)

    for c_idx in range(1, len(output_cols) + 1):
        hcell = ws.cell(row=1, column=c_idx)
        hcell.font = header_font
        hcell.fill = header_fill
        hcell.alignment = Alignment(horizontal="center", vertical="center")
        hcell.border = thin_border

    d_list = df["序号D"].tolist()
    d_count_map = {}
    for d_val in d_list:
        key = excel_text_str(d_val)
        d_count_map[key] = d_count_map.get(key, 0) + 1
    n = len(d_list)
    start = 0
    while start < n:
        end = start
        while end + 1 < n and d_list[end + 1] == d_list[start]:
            end += 1
        d_key = excel_text_str(d_list[start])
        if d_count_map.get(d_key, 0) > 1:
            top_cell = ws.cell(row=start + 2, column=4)
            top_cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")
            top_cell.font = Font(name=body_font.name, size=body_font.size, bold=True)
        if end > start:
            ws.merge_cells(
                start_row=start + 2,
                start_column=4,
                end_row=end + 2,
                end_column=4,
            )
        start = end + 1

    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = 48
    ws.print_options.horizontalCentered = True

    wb.save(output)
    return output.getvalue()


def generate_rename_list(df):
    """Build image-rename list from final print list: B→原图片名称, A→新图片名称, Q→分类备注."""
    rename_df = pd.DataFrame(
        {
            "原图片名称": df["purchase-date"].map(excel_text_str),
            "新图片名称": df["运单号"].map(excel_text_str),
            "分类备注": df["图片名称"].map(excel_text_str),
        }
    )
    output = io.BytesIO()
    rename_df.to_excel(output, index=False, engine="openpyxl")
    return output.getvalue()


PRINT_LIST_COL_SEQ_D = "序号D"
PRINT_LIST_COL_MATERIAL = "材质"
PRINT_LIST_COL_SIZE = "尺寸"
PRINT_LIST_EXCEL_D = "D"
PRINT_LIST_EXCEL_G = "G"
PRINT_LIST_EXCEL_H = "H"

SHIPPING_HEADER = [
    "*仓库编码",
    "*国家编码",
    "*渠道编码",
    "参考编号1",
    "参考编号2",
    "参考编号3",
    "签收服务",
    "*是否是FBA",
    "FBA仓库编码",
    "收件人联系人",
    "收件人公司",
    "收件人邮编",
    "收件人地址第一行",
    "收件人地址第二行",
    "收件人城市",
    "收件人州",
    "收件人电话",
    "收件人Email",
    "发件人联系人",
    "发件人公司",
    "发件人地址第一行",
    "发件人地址第二行",
    "发件人城市",
    "发件人身份证号",
    "发件人州",
    "发件人国家编码",
    "发件人邮编",
    "发件人电话",
    "发件人Email",
    "发件人税号",
    "申报币种",
    "尺寸单位",
    "预报重量单位",
    "*箱数",
    "*重量(KG)",
    "*长(CM)",
    "*宽(CM)",
    "*高(CM)",
    "申报中文.1",
    "申报英文.1",
    "数量.1",
    "价值.1",
    "申报重量.1",
    "产品SKU.1",
]

SHIPPING_WEIGHT_MAP = {
    "810": {"canvas": (0.15, 0.1, 0.15), "non_canvas": (0.4, 0.1, 0.4)},
    "1114": {"canvas": (0.2, 0.1, 0.2), "non_canvas": (0.5, 0.1, 0.5)},
    "1117": {"canvas": (0.22, 0.12, 0.22), "non_canvas": (0.52, 0.12, 0.52)},
    "1212": {"canvas": (0.22, 0.12, 0.22), "non_canvas": (0.52, 0.12, 0.52)},
    "1216": {"canvas": (0.26, 0.14, 0.26), "non_canvas": (0.56, 0.14, 0.56)},
    "1218": {"canvas": (0.22, 0.15, 0.42), "non_canvas": (0.5, 0.15, 0.5)},
    "1236": {"canvas": (0.22, 0.16, 0.43), "non_canvas": (0.5, 0.16, 0.5)},
    "1313": {"canvas": (0.25, 0.18, 0.44), "non_canvas": (0.5, 0.18, 0.5)},
    "1620": {"canvas": (0.2, 0.18, 0.45), "non_canvas": (0.55, 0.18, 0.55)},
    "1624": {"canvas": (0.22, 0.18, 0.46), "non_canvas": (0.6, 0.18, 0.6)},
    "1824": {"canvas": (0.24, 0.18, 0.6), "non_canvas": (0.74, 0.18, 0.74)},
    "2020": {"canvas": (0.26, 0.18, 0.65), "non_canvas": (0.75, 0.18, 0.75)},
    "2024": {"canvas": (0.28, 0.18, 0.66), "non_canvas": (0.78, 0.18, 0.78)},
    "2030": {"canvas": (0.34, 0.18, 0.8), "non_canvas": (0.85, 0.18, 0.85)},
    "2040": {"canvas": (0.4, 0.18, 0.82), "non_canvas": (0.9, 0.18, 0.9)},
    "2228": {"canvas": (0.38, 0.18, 0.95), "non_canvas": (0.95, 0.18, 0.95)},
    "2234": {"canvas": (0.4, 0.18, 0.98), "non_canvas": (0.98, 0.18, 0.98)},
    "2430": {"canvas": (0.38, 0.18, 0.75), "non_canvas": (0.88, 0.18, 0.88)},
    "2436": {"canvas": (0.42, 0.2, 0.95), "non_canvas": (0.98, 0.2, 0.98)},
    "2740": {"canvas": (0.5, 0.2, 1.1), "non_canvas": (1.2, 0.2, 1.2)},
    "27.5x39": {"canvas": (0.52, 0.2, 1.1), "non_canvas": (1.25, 0.2, 1.25)},
    "3030": {"canvas": (0.46, 0.2, 1.15), "non_canvas": (1.3, 0.2, 1.3)},
    "3040": {"canvas": (0.56, 0.22, 1.16), "non_canvas": (1.35, 0.2, 1.35)},
    "3045": {"canvas": (0.6, 0.22, 1.18), "non_canvas": (1.45, 0.2, 1.45)},
    "3648": {"canvas": (0.74, 0.3, 1.7), "non_canvas": (1.7, 0.3, 1.7)},
    "3650": {"canvas": (0.76, 0.3, 1.75), "non_canvas": (1.75, 0.3, 1.75)},
    "3672": {"canvas": (0.98, 0.35, 1.8), "non_canvas": (1.95, 0.35, 1.95)},
    "4040": {"canvas": (0.72, 0.35, 1.65), "non_canvas": (1.95, 0.35, 1.95)},
    "4048": {"canvas": (0.8, 0.4, 1.55), "non_canvas": (1.98, 0.4, 1.98)},
    "4060": {"canvas": (0.94, 0.5, 1.65), "non_canvas": (2.0, 0.5, 2.0)},
    "4545": {"canvas": (1.16, 0.55, 2.5), "non_canvas": (3.0, 0.55, 3.0)},
}


def _normalize_date_part(date_part):
    """Normalize title date, e.g. '6.9' -> '0609', '4.20' -> '0420'."""
    date_part = (date_part or "").strip()
    if not date_part:
        return ""
    if "." in date_part:
        month, day = date_part.split(".", 1)
        return month.zfill(2) + day.zfill(2)
    date_str = date_part.replace(".", "")
    return date_str.zfill(4)


def _parse_pdf_title(pdf_title):
    """Extract short code and normalized date from PDF title, e.g. '4.20_Order List_XM'."""
    title = (pdf_title or "").strip()
    if not title:
        return "XM", ""
    parts = title.split("_")
    if len(parts) >= 2:
        return parts[-1], _normalize_date_part(parts[0])
    return title, _normalize_date_part(title)


def _is_valid_seq_d(val):
    text = excel_text_str(val)
    return bool(text) and text != "0"


def _size_to_numeric(size_val):
    try:
        return int(float(str(size_val).strip()))
    except Exception:
        return 0


def _validate_print_list_columns(df):
    """Ensure final print-list columns match Excel D/G/H mapping before shipping export."""
    required = [
        PRINT_LIST_COL_SEQ_D,
        PRINT_LIST_COL_MATERIAL,
        PRINT_LIST_COL_SIZE,
        "purchase-date",
        "姓名",
        "邮编",
        "地址一",
        "地址二",
        "城市",
        "州",
        "电话",
    ]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            f"最终打印清单缺少必要列: {missing}；当前列名为: {list(df.columns)}"
        )


def _read_print_list_value(row, col_name):
    """Read one print-list field by stable column name (never positional index)."""
    if col_name not in row.index:
        raise KeyError(
            f"列 '{col_name}' 不存在于最终打印清单，当前列名为: {list(row.index)}"
        )
    return row[col_name]


def _add_excel_column_aliases(df):
    """Expose Excel D/G/H aliases on merged print-list rows for AJ calculation."""
    aliased = df.copy()
    aliased[PRINT_LIST_EXCEL_D] = aliased[PRINT_LIST_COL_SEQ_D]
    aliased[PRINT_LIST_EXCEL_G] = aliased[PRINT_LIST_COL_MATERIAL]
    aliased[PRINT_LIST_EXCEL_H] = aliased[PRINT_LIST_COL_SIZE]
    return aliased


def _merge_print_list_by_seq_d(df):
    """Dedupe by D column: keep first row; attach all G/H pairs for multi-piece AJ calc."""
    work = df.loc[df[PRINT_LIST_COL_SEQ_D].apply(_is_valid_seq_d)].copy()
    merged_rows = []

    for _, grp in work.groupby(PRINT_LIST_COL_SEQ_D, sort=False):
        rep = grp.iloc[0].copy()
        rep["_group_count"] = len(grp)
        rep["_group_g_list"] = grp[PRINT_LIST_COL_MATERIAL].tolist()
        rep["_group_h_list"] = grp[PRINT_LIST_COL_SIZE].tolist()
        rep["_group_qty_list"] = grp["数量"].tolist()
        merged_rows.append(rep)

    merged = pd.DataFrame(merged_rows).reset_index(drop=True)
    return _add_excel_column_aliases(merged)


def get_inch_val(g_val, h_val):
    """Convert one print-list row to inch baseline (no +1). Canvas uses H; non-Canvas uses G."""
    if "Canvas" in str(g_val):
        h_str = str(abs(_size_to_numeric(h_val)))
        if len(h_str) >= 4:
            return int(h_str[:2])
        if len(h_str) == 3:
            return int(h_str[0])
        return 0

    try:
        match = re.search(r"(\d+)", str(g_val))
        if not match:
            return 0
        num = float(match.group(1))
        return int(round(num / 2.54))
    except Exception:
        return 0


def calculate_aj(row):
    """Compute *长(CM): max inch across all D-group rows, then +1."""
    g_list = row["_group_g_list"]
    h_list = row["_group_h_list"]
    group_count = int(row.get("_group_count", 1))

    inch_vals = [get_inch_val(g_val, h_val) for g_val, h_val in zip(g_list, h_list)]
    print(
        f"DEBUG AJ D组 group_count={group_count}, "
        f"G列表={g_list}, H列表={h_list}, 英寸值={inch_vals}"
    )

    max_inch = max(inch_vals) if inch_vals else 0
    return max_inch + 1


def _normalize_weight_size_key(h_val):
    """Normalize print-list size for WEIGHT_MAP lookup."""
    size_num = _size_to_numeric(h_val)
    if size_num:
        return str(abs(size_num))
    return str(h_val).replace(" ", "").lower()


def _calculate_ai_weight(print_row):
    """Precise shipping weight: per-size formula with increment and min threshold."""
    g_list = print_row.get("_group_g_list", [])
    h_list = print_row.get("_group_h_list", [])
    qty_list = print_row.get("_group_qty_list", [])

    total_items = 0
    for q in qty_list:
        try:
            total_items += int(float(str(q).strip()))
        except Exception:
            total_items += 1

    if total_items < 1:
        total_items = 1

    calculated_weights = []
    for g_val, h_val in zip(g_list, h_list):
        clean_h = _normalize_weight_size_key(h_val)
        is_canvas = "canvas" in str(g_val).lower()

        if clean_h in SHIPPING_WEIGHT_MAP:
            base, inc, min_val = (
                SHIPPING_WEIGHT_MAP[clean_h]["canvas"]
                if is_canvas
                else SHIPPING_WEIGHT_MAP[clean_h]["non_canvas"]
            )
            item_total_weight = base + (total_items - 1) * inc
            final_weight = max(item_total_weight, min_val)
        else:
            final_weight = 1.0 + (total_items - 1) * 0.2

        calculated_weights.append(final_weight)

    if not calculated_weights:
        return "1"

    total_weight = round(max(calculated_weights), 2)
    return f"{total_weight:g}"


def _empty_shipping_row():
    return {col: "" for col in SHIPPING_HEADER}


def _build_shipping_data_row(print_row, short_code, date_str):
    """Map legacy A-AQ fill rules onto named shipping-template columns."""
    row = _empty_shipping_row()

    ai_val = _calculate_ai_weight(print_row)
    aj_val = print_row["_aj_value"]
    c_val = "OnTrac" if float(aj_val) > 30 else "AMAZON_SHIPPING-GROUND_1_6"

    seq_d = excel_text_str(_read_print_list_value(print_row, PRINT_LIST_COL_SEQ_D))
    order_no = excel_text_str(_read_print_list_value(print_row, "purchase-date"))
    name = excel_text_str(_read_print_list_value(print_row, "姓名"))
    ref_id = f"{short_code}-{date_str}-{seq_d}"

    row["*仓库编码"] = "PRINKO-SD"
    row["*国家编码"] = "US"
    row["*渠道编码"] = c_val
    row["参考编号1"] = f"{ref_id}-{order_no}"
    row["签收服务"] = "No"
    row["*是否是FBA"] = "FALSE"
    row["收件人联系人"] = f"{ref_id}-{name}"
    row["收件人邮编"] = excel_text_str(_read_print_list_value(print_row, "邮编"))
    row["收件人地址第一行"] = excel_text_str(_read_print_list_value(print_row, "地址一"))
    row["收件人地址第二行"] = excel_text_str(_read_print_list_value(print_row, "地址二"))
    row["收件人城市"] = excel_text_str(_read_print_list_value(print_row, "城市"))
    row["收件人州"] = excel_text_str(_read_print_list_value(print_row, "州"))
    row["收件人电话"] = excel_text_str(_read_print_list_value(print_row, "电话"))
    row["发件人联系人"] = "JM"
    row["发件人地址第一行"] = "14207 Monte Vista Ave"
    row["发件人城市"] = "Chino Hills"
    row["发件人州"] = "CA"
    row["发件人邮编"] = "91180"
    row["发件人电话"] = "+1 619-854-2705"
    row["尺寸单位"] = "inch"
    row["预报重量单位"] = "lb"
    row["*箱数"] = "1"
    row["*重量(KG)"] = ai_val
    row["*长(CM)"] = aj_val
    row["*宽(CM)"] = "5"
    row["*高(CM)"] = "1"
    row["申报中文.1"] = "定制装饰画"
    row["申报英文.1"] = "Customer Canvas"
    row["数量.1"] = "1"
    row["价值.1"] = "35"
    row["申报重量.1"] = ai_val
    return row


def generate_shipping_list(df, pdf_title):
    """Build upload shipping-number DataFrame with Chinese headers and deduplicated orders."""
    print("最终打印清单 columns:", list(df.columns))
    _validate_print_list_columns(df)

    short_code, date_str = _parse_pdf_title(pdf_title)
    merged_df = _merge_print_list_by_seq_d(df)
    merged_df["_aj_value"] = merged_df.apply(calculate_aj, axis=1)

    shipping_rows = merged_df.apply(
        lambda print_row: _build_shipping_data_row(print_row, short_code, date_str),
        axis=1,
    )
    shipping_df = pd.DataFrame(shipping_rows.tolist(), columns=SHIPPING_HEADER)
    shipping_df["签收服务"] = "No"
    shipping_df["*是否是FBA"] = "FALSE"
    return shipping_df


def export_shipping_list_excel(shipping_df):
    """Export shipping-list DataFrame to Excel bytes with Chinese header row."""
    output = io.BytesIO()
    shipping_df.to_excel(output, index=False, header=True, engine="openpyxl")
    return output.getvalue()
